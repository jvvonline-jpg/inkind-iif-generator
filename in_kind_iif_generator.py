#!/usr/bin/env python3
"""
Cornerstones, Inc. — In-Kind Gifts → QuickBooks IIF Generator
=============================================================

Reads the quarterly in-kind gift Excel workbooks from a folder and
produces a QuickBooks-ready .iif file for the general journal entry.

Input folder must contain:
  - All Gifts <Q#> FY<YY> REV<NN>.xlsx   (master list of all gifts)
  - ERCS Gifts <Q#> FY<YY> REV<NN>.xlsx  (with 2 sheets that split ERCS)
  - (optional) ESHP / Food Hub / General CS / HNRC / Hypothermia / LLC workbooks
    — these are for reference; totals come from the All Gifts file.

Output: one .iif file with a single General Journal transaction where
each program area produces a Debit (In Kind - Programs, 9026.1) and
an equal Credit (In-Kind Goods & Prof Services, 8111).

Usage (GUI):   python3 in_kind_iif_generator.py
Usage (CLI):   python3 in_kind_iif_generator.py --folder "/path/to/folder" \
                    --posting-date 12/31/2025 --quarter "Q2 FY26" \
                    --out "/path/to/In-Kind Q2 IIF.iif"
"""

from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("openpyxl is required.  Install with:  pip3 install openpyxl --break-system-packages")
    raise


# ---------------------------------------------------------------------------
# Mapping tables — Cornerstones, Inc. chart-of-accounts specific
# ---------------------------------------------------------------------------

DEBIT_ACCOUNT  = "DIRECT PROGRAM EXPENSES:IN-KIND Program Expense:In Kind - Programs"
CREDIT_ACCOUNT = "In-Kind Contributions/Donations:In Kind Goods & Prof. Services"

# The full class path as it exists in QuickBooks.
# Keyed by the short class number used in the Template.
CLASS_PATHS = {
    "41000": "40000 CB & NR:41000 ASAPP",
    "42000": "40000 CB & NR:42000 HNRC",
    "43000": "40000 CB & NR:43000 Community Connected Sites",
    "51000": "50000 Affrdble Hsing Prtnrshps:51000 Scattered Site Admin",
    "81000": "80000 Community Programs:81000 FREE Food Hub",
    "32100": "30000 Housing & Community Serv:32000 Shelter:32100 Shelter Ops & Outreach",
    "32200": "30000 Housing & Community Serv:32000 Shelter:32200 Hypothermia",
    "32302": "30000 Housing & Community Serv:32000 Shelter:32300 Shelter CM & Aftercare:32302 OPEH HH W Child Cntrct",
    "32303": "30000 Housing & Community Serv:32000 Shelter:32300 Shelter CM & Aftercare:32303 OPEH HH WO Child Cntrc",
    "21000": "20000 Resource Development:21000 Fundraising",
    "61000": "60000 Family Self-Sufficiency:61000 Laurel Learning Center",
}

# All Gifts file's Fund Description → (class, memo suffix) mapping.
# ERCS is handled separately because it splits across 3 classes.
FUND_TO_CLASS = {
    "Assistance Services & Pantry Program":                           ("41000", "ASAPP"),
    "Community Connected Sites":                                      ("43000", "Community Connected Sites"),
    "Emergency Support Housing Program":                              ("51000", "ESHP (HOUSE)"),
    "Food Hub":                                                       ("81000", "Food Hub"),
    "Hypothermia":                                                    ("32200", "Hypo - TOS"),
    "General Cornerstones Fund":                                      ("21000", "General CS"),
    "Herndon Neighborhood Resource Center Operations & Adult Services": ("42000", "HNRC -OAS"),
    "Laurel Learning Center":                                         ("61000", "LLC"),
    # ERCS (Embry Rucker Community Shelter) handled specially.
}

# Order the lines appear in the JE — matches Jon's Template Q2 FY26 layout
LINE_ORDER = [
    "ASAPP", "Community Connected Sites", "ESHP (HOUSE)", "Food Hub",
    "ERCS Shelter", "Hypo - TOS", "HH w/ Children", "HH w/o Children",
    "General CS", "HNRC -OAS", "LLC",
]

LINE_TO_CLASS = {
    "ASAPP":                      "41000",
    "Community Connected Sites":  "43000",
    "ESHP (HOUSE)":               "51000",
    "Food Hub":                   "81000",
    "ERCS Shelter":               "32100",
    "Hypo - TOS":                 "32200",
    "HH w/ Children":             "32302",
    "HH w/o Children":            "32303",
    "General CS":                 "21000",
    "HNRC -OAS":                  "42000",
    "LLC":                        "61000",
}


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

@dataclass
class LineItem:
    label: str        # short program name used in memo, e.g. "ASAPP"
    class_num: str    # e.g. "41000"
    amount: float     # debit amount (credit is the negation)


def find_workbook(folder: Path, prefix: str) -> Optional[Path]:
    """Return the first .xlsx matching the given prefix (case-insensitive)."""
    prefix_low = prefix.lower()
    for p in sorted(folder.iterdir()):
        if p.suffix.lower() == ".xlsx" and p.name.lower().startswith(prefix_low):
            return p
    return None


def read_all_gifts(all_gifts_path: Path) -> dict[str, float]:
    """Sum Gift Amount by Fund Description.  Returns {fund_description: total}."""
    wb = load_workbook(all_gifts_path, data_only=True, read_only=True)
    ws = wb.active
    totals: dict[str, float] = defaultdict(float)
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            hdr = [str(c).strip() if c else "" for c in row]
            continue
        if all(v is None for v in row):
            continue
        # Tolerate small column drifts by looking up header positions
        try:
            fund_idx   = hdr.index("Fund Description")
            amount_idx = hdr.index("Gift Amount")
        except ValueError:
            fund_idx, amount_idx = 4, 6  # fall back to sample-file layout
        fund = row[fund_idx]
        amt  = row[amount_idx]
        if fund and amt is not None:
            totals[str(fund).strip()] += float(amt)
    wb.close()
    return dict(totals)


def read_ercs_split(ercs_path: Path) -> tuple[float, float]:
    """
    Return (shelter_total, hh_total) from the ERCS workbook.
      - 'Shelter' sheet name starts with "351" (e.g. "351.12")        → ERCS Shelter class 32100
      - 'HH'      sheet name starts with "381" (contains "381" & "385") → split 50/50 into 32302/32303
    """
    wb = load_workbook(ercs_path, data_only=True, read_only=True)
    shelter = hh = 0.0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total = 0.0
        hdr = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                hdr = [str(c).strip() if c else "" for c in row]
                continue
            if all(v is None for v in row):
                continue
            # skip the "Totals:" row (col A text)
            if row[0] and isinstance(row[0], str) and row[0].strip().lower().startswith("total"):
                continue
            try:
                amount_idx = hdr.index("Gift Amount")
            except ValueError:
                amount_idx = 6
            amt = row[amount_idx]
            if amt is not None:
                total += float(amt)
        key = sheet_name.strip()
        if key.startswith("351"):
            shelter = total
        elif key.startswith("381") or "381" in key:
            hh = total
    wb.close()
    return shelter, hh


def build_line_items(all_gifts_path: Path, ercs_path: Path) -> list[LineItem]:
    totals = read_all_gifts(all_gifts_path)
    shelter, hh = read_ercs_split(ercs_path)

    # Build map: short label → amount
    amounts: dict[str, float] = defaultdict(float)
    for fund, amt in totals.items():
        if fund in FUND_TO_CLASS:
            _, label = FUND_TO_CLASS[fund]
            amounts[label] += amt
        elif fund == "Embry Rucker Community Shelter":
            # ERCS is split across 3 lines, handled below
            pass
        else:
            # Unknown fund — treat as a new line so nothing is lost
            amounts[fund] += amt
    amounts["ERCS Shelter"]    = shelter
    amounts["HH w/ Children"]  = hh / 2.0
    amounts["HH w/o Children"] = hh / 2.0

    # Cross-check: sum of ERCS parts should match ERCS total in All Gifts
    ercs_total_from_all = totals.get("Embry Rucker Community Shelter", 0.0)
    ercs_total_from_split = shelter + hh
    if abs(ercs_total_from_all - ercs_total_from_split) > 0.01:
        print(
            f"[WARN] ERCS total in All Gifts ({ercs_total_from_all:,.2f}) does not match "
            f"sum of ERCS workbook sheets ({ercs_total_from_split:,.2f}).  "
            "Check that both workbooks are from the same period.",
            file=sys.stderr,
        )

    items: list[LineItem] = []
    for label in LINE_ORDER:
        amt = round(amounts.get(label, 0.0), 2)
        if amt == 0:
            # Omit zero-value lines from the IIF (matches sample behavior)
            continue
        items.append(LineItem(label=label, class_num=LINE_TO_CLASS[label], amount=amt))
    # Any unexpected labels
    for label, amt in amounts.items():
        if label not in LINE_TO_CLASS and round(amt, 2) != 0:
            print(f"[WARN] Unmapped line '{label}' with amount {amt:,.2f} — skipped.", file=sys.stderr)
    return items


# ---------------------------------------------------------------------------
# IIF writer
# ---------------------------------------------------------------------------

IIF_HEADER = (
    "!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\n"
    "!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\n"
    "!ENDTRNS\t\t\t\t\t\t \t\t\t\n"
)


def _fmt_amount(x: float) -> str:
    """Match sample IIF: no fixed decimals, trailing zeros stripped."""
    s = f"{x:.3f}".rstrip("0").rstrip(".")
    return s if s else "0"


def build_iif(items: list[LineItem], posting_date: str, quarter: str, docnum: str = "In-Kind") -> str:
    out = [IIF_HEADER]
    first = True
    qtr_short = quarter.split()[0] if quarter else "Q"   # e.g. "Q2"
    for item in items:
        class_path = CLASS_PATHS[item.class_num]
        exp_memo = f"{quarter} in-kind program exp per {qtr_short} summary - {item.label}"
        rev_memo = f"{quarter} in-kind program rev per {qtr_short} summary - {item.label}"
        # Debit line
        tag = "TRNS" if first else "SPL"
        out.append(
            f"{tag}\t\tGENERAL JOURNAL\t{posting_date}\t{DEBIT_ACCOUNT}\t\t{class_path}\t"
            f"{_fmt_amount(item.amount)}\t{docnum if first else ''}\t{exp_memo}\n"
        )
        # Credit line (always SPL)
        out.append(
            f"SPL\t\tGENERAL JOURNAL\t{posting_date}\t{CREDIT_ACCOUNT}\t\t{class_path}\t"
            f"{_fmt_amount(-item.amount)}\t\t{rev_memo}\n"
        )
        first = False
    out.append("ENDTRNS\n")
    return "".join(out)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def generate(folder: Path, out_path: Path, posting_date: str, quarter: str) -> tuple[list[LineItem], str]:
    all_gifts = find_workbook(folder, "All Gifts")
    ercs      = find_workbook(folder, "ERCS Gifts")
    if not all_gifts:
        raise FileNotFoundError(f"Could not find 'All Gifts ....xlsx' in {folder}")
    if not ercs:
        raise FileNotFoundError(f"Could not find 'ERCS Gifts ....xlsx' in {folder}")

    items = build_line_items(all_gifts, ercs)
    iif = build_iif(items, posting_date=posting_date, quarter=quarter)
    out_path.write_text(iif, encoding="utf-8")
    return items, iif


# ---------------------------------------------------------------------------
# CLI / GUI entry point
# ---------------------------------------------------------------------------

def _cli() -> None:
    ap = argparse.ArgumentParser(description="Generate Cornerstones In-Kind JE IIF file")
    ap.add_argument("--folder", required=False, help="Folder containing the quarterly gift workbooks")
    ap.add_argument("--posting-date", default="12/31/2025", help='MM/DD/YYYY (default 12/31/2025)')
    ap.add_argument("--quarter", default="Q2 FY26", help='Quarter label used in memos (e.g. "Q2 FY26")')
    ap.add_argument("--out", help="Path to output .iif file")
    ap.add_argument("--gui", action="store_true", help="Launch the GUI")
    args = ap.parse_args()

    if args.gui or not args.folder:
        _gui()
        return

    folder = Path(args.folder).expanduser().resolve()
    out = Path(args.out).expanduser().resolve() if args.out else folder / f"In-Kind {args.quarter.split()[0]} IIF.iif"
    items, _ = generate(folder, out, args.posting_date, args.quarter)
    total = sum(i.amount for i in items)
    print(f"Wrote {out}")
    for i in items:
        print(f"  {i.label:20s} {i.class_num}  {i.amount:>14,.2f}")
    print(f"  {'TOTAL':20s}         {total:>14,.2f}")


def _gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("Cornerstones · In-Kind JE → IIF")
    root.geometry("640x440")

    # State
    folder_var   = tk.StringVar()
    date_var     = tk.StringVar(value="12/31/2025")
    quarter_var  = tk.StringVar(value="Q2 FY26")
    out_var      = tk.StringVar()

    frm = ttk.Frame(root, padding=14)
    frm.pack(fill="both", expand=True)

    def _browse_folder():
        d = filedialog.askdirectory(title="Select the quarterly gifts folder")
        if d:
            folder_var.set(d)
            # Default output file
            q = quarter_var.get().strip().split()[0] or "Q"
            out_var.set(os.path.join(d, f"In-Kind {q} IIF.iif"))

    def _browse_out():
        f = filedialog.asksaveasfilename(
            defaultextension=".iif",
            filetypes=[("QuickBooks IIF", "*.iif")],
            initialfile=os.path.basename(out_var.get()) or "In-Kind.iif",
        )
        if f:
            out_var.set(f)

    ttk.Label(frm, text="Input folder:").grid(row=0, column=0, sticky="w")
    ttk.Entry(frm, textvariable=folder_var, width=60).grid(row=0, column=1, sticky="we", padx=6)
    ttk.Button(frm, text="Browse…", command=_browse_folder).grid(row=0, column=2)

    ttk.Label(frm, text="Posting date (MM/DD/YYYY):").grid(row=1, column=0, sticky="w", pady=(12, 0))
    ttk.Entry(frm, textvariable=date_var, width=18).grid(row=1, column=1, sticky="w", pady=(12, 0), padx=6)

    ttk.Label(frm, text="Quarter label (for memos):").grid(row=2, column=0, sticky="w", pady=(8, 0))
    ttk.Entry(frm, textvariable=quarter_var, width=18).grid(row=2, column=1, sticky="w", pady=(8, 0), padx=6)

    ttk.Label(frm, text="Output .iif file:").grid(row=3, column=0, sticky="w", pady=(8, 0))
    ttk.Entry(frm, textvariable=out_var, width=60).grid(row=3, column=1, sticky="we", pady=(8, 0), padx=6)
    ttk.Button(frm, text="Save as…", command=_browse_out).grid(row=3, column=2, pady=(8, 0))

    # Results / log area
    txt = tk.Text(frm, height=14, width=78, wrap="none")
    txt.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(12, 0))
    scr = ttk.Scrollbar(frm, orient="vertical", command=txt.yview)
    scr.grid(row=5, column=3, sticky="ns", pady=(12, 0))
    txt.configure(yscrollcommand=scr.set)

    frm.columnconfigure(1, weight=1)
    frm.rowconfigure(5, weight=1)

    def _run():
        txt.delete("1.0", "end")
        folder = folder_var.get().strip()
        out    = out_var.get().strip()
        date   = date_var.get().strip()
        qtr    = quarter_var.get().strip()
        if not folder or not out:
            messagebox.showerror("Missing info", "Please choose an input folder and output file.")
            return
        try:
            items, iif = generate(Path(folder), Path(out), date, qtr)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        total = sum(i.amount for i in items)
        txt.insert("end", f"✓ Wrote {out}\n\n")
        txt.insert("end", f"{'Program':22s} {'Class':8s} {'Amount':>14s}\n")
        txt.insert("end", "-" * 50 + "\n")
        for i in items:
            txt.insert("end", f"{i.label:22s} {i.class_num:8s} {i.amount:>14,.2f}\n")
        txt.insert("end", "-" * 50 + "\n")
        txt.insert("end", f"{'TOTAL':22s} {'':8s} {total:>14,.2f}\n")

    ttk.Button(frm, text="Generate IIF", command=_run).grid(row=4, column=0, columnspan=3, pady=12)

    root.mainloop()


if __name__ == "__main__":
    _cli()
