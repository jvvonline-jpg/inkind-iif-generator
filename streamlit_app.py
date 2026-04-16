#!/usr/bin/env python3
"""
Cornerstones, Inc. — In-Kind Gifts → QuickBooks IIF Generator
=============================================================

Reads the quarterly in-kind gift Excel workbooks and produces a
QuickBooks-ready .iif file for the general journal entry.

Input files required:
  - All Gifts <Q#> FY<YY> REV<NN>.xlsx   (master list of all gifts)
  - ERCS Gifts <Q#> FY<YY> REV<NN>.xlsx  (with 2 sheets that split ERCS)

Output: one .iif file with a single General Journal transaction where
each program area produces a Debit (In Kind - Programs, 9026.1) and
an equal Credit (In-Kind Goods & Prof Services, 8111).

Runs as a Streamlit web app.
"""

from __future__ import annotations

import io
import sys
from collections import defaultdict
from dataclasses import dataclass
from typing import Optional

import streamlit as st
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Mapping tables — Cornerstones, Inc. chart-of-accounts specific
# ---------------------------------------------------------------------------

DEBIT_ACCOUNT  = "DIRECT PROGRAM EXPENSES:IN-KIND Program Expense:In Kind - Programs"
CREDIT_ACCOUNT = "In-Kind Contributions/Donations:In Kind Goods & Prof. Services"

# The full class path as it exists in QuickBooks.
# Keyed by the short class number used in the Template.
CLASS_PATHS = {
    "41000": "40000 CB & NR:41000 ASAPP",
    "42000": "40000 CB & NR:42000 HNRC",
    "43000": "40000 CB & NR:43000 Community Connected Sites",
    "51000": "50000 Affrdble Hsing Prtnrshps:51000 Scattered Site Admin",
    "81000": "80000 Community Programs:81000 FREE Food Hub",
    "32100": "30000 Housing & Community Serv:32000 Shelter:32100 Shelter Ops & Outreach",
    "32200": "30000 Housing & Community Serv:32000 Shelter:32200 Hypothermia",
    "32302": "30000 Housing & Community Serv:32000 Shelter:32300 Shelter CM & Aftercare:32302 OPEH HH W Child Cntrct",
    "32303": "30000 Housing & Community Serv:32000 Shelter:32300 Shelter CM & Aftercare:32303 OPEH HH WO Child Cntrc",
    "21000": "20000 Resource Development:21000 Fundraising",
    "61000": "60000 Family Self-Sufficiency:61000 Laurel Learning Center",
    "33000": "30000 Housing & Community Serv:33000 Supportive Housing",
    "45000": "40000 CB & NR:45000 Opportunity Neighborhoods",
    "63000": "60000 Family Self-Sufficiency:63000 Connections 4 Hope",
}

# All Gifts file's Fund Description → (class, memo suffix) mapping.
# ERCS is handled separately because it splits across 3 classes.
FUND_TO_CLASS = {
    "Assistance Services & Pantry Program":                           ("41000", "ASAPP"),
    "Community Connected Sites":                                      ("43000", "Community Connected Sites"),
    "Emergency Support Housing Program":                              ("51000", "ESHP (HOUSE)"),
    "Food Hub":                                                       ("81000", "Food Hub"),
    "Hypothermia":                                                    ("32200", "Hypo - TOS"),
    "General Cornerstones Fund":                                      ("21000", "General CS"),
    "Herndon Neighborhood Resource Center Operations & Adult Services": ("42000", "HNRC -OAS"),
    "Laurel Learning Center":                                         ("61000", "LLC"),
    "Supportive Housing":                                             ("33000", "Supportive Housing"),
    "Opportunity Neighborhood (RestON)":                              ("45000", "ON (RestON)"),
    "C4HP CCFP Svs Contract":                                        ("63000", "C4HP"),
    # ERCS (Embry Rucker Community Shelter) handled specially.
}

# Order the lines appear in the JE — matches the Template layout
LINE_ORDER = [
    "ASAPP", "Community Connected Sites", "ESHP (HOUSE)", "Food Hub",
    "ERCS Shelter", "Hypo - TOS", "HH w/ Children", "HH w/o Children",
    "General CS", "HNRC -OAS", "LLC",
    "Supportive Housing", "ON (RestON)", "C4HP",
]

LINE_TO_CLASS = {
    "ASAPP":                      "41000",
    "Community Connected Sites":  "43000",
    "ESHP (HOUSE)":               "51000",
    "Food Hub":                   "81000",
    "ERCS Shelter":               "32100",
    "Hypo - TOS":                 "32200",
    "HH w/ Children":             "32302",
    "HH w/o Children":            "32303",
    "General CS":                 "21000",
    "HNRC -OAS":                  "42000",
    "LLC":                        "61000",
    "Supportive Housing":         "33000",
    "ON (RestON)":                "45000",
    "C4HP":                       "63000",
}


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

@dataclass
class LineItem:
    label: str        # short program name used in memo, e.g. "ASAPP"
    class_num: str    # e.g. "41000"
    amount: float     # debit amount (credit is the negation)


def read_all_gifts(file_bytes: bytes) -> dict[str, float]:
    """Sum Gift Amount by Fund Description.  Returns {fund_description: total}."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    totals: dict[str, float] = defaultdict(float)
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            hdr = [str(c).strip() if c else "" for c in row]
            continue
        if all(v is None for v in row):
            continue
        try:
            fund_idx   = hdr.index("Fund Description")
            amount_idx = hdr.index("Gift Amount")
        except ValueError:
            fund_idx, amount_idx = 4, 6  # fall back to sample-file layout
        fund = row[fund_idx]
        amt  = row[amount_idx]
        if fund and amt is not None:
            totals[str(fund).strip()] += float(amt)
    wb.close()
    return dict(totals)


def read_ercs_split(file_bytes: bytes) -> tuple[float, float]:
    """
    Return (shelter_total, hh_total) from the ERCS workbook.
      - Sheet starting with "351" → ERCS Shelter class 32100
      - Sheet starting with "381" → split 50/50 into 32302/32303
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    shelter = hh = 0.0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total = 0.0
        hdr = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                hdr = [str(c).strip() if c else "" for c in row]
                continue
            if all(v is None for v in row):
                continue
            # skip the "Totals:" row
            if row[0] and isinstance(row[0], str) and row[0].strip().lower().startswith("total"):
                continue
            try:
                amount_idx = hdr.index("Gift Amount")
            except ValueError:
                amount_idx = 6
            amt = row[amount_idx]
            if amt is not None:
                total += float(amt)
        key = sheet_name.strip()
        if key.startswith("351"):
            shelter = total
        elif key.startswith("381") or "381" in key:
            hh = total
    wb.close()
    return shelter, hh


def build_line_items(all_gifts_bytes: bytes, ercs_bytes: bytes) -> tuple[list[LineItem], list[str]]:
    """Build line items and return (items, warnings)."""
    warnings = []
    totals = read_all_gifts(all_gifts_bytes)
    shelter, hh = read_ercs_split(ercs_bytes)

    amounts: dict[str, float] = defaultdict(float)
    for fund, amt in totals.items():
        if fund in FUND_TO_CLASS:
            _, label = FUND_TO_CLASS[fund]
            amounts[label] += amt
        elif fund == "Embry Rucker Community Shelter":
            pass  # handled below via ERCS split
        else:
            amounts[fund] += amt
    amounts["ERCS Shelter"]    = shelter
    amounts["HH w/ Children"]  = hh / 2.0
    amounts["HH w/o Children"] = hh / 2.0

    # Cross-check
    ercs_total_from_all = totals.get("Embry Rucker Community Shelter", 0.0)
    ercs_total_from_split = shelter + hh
    if abs(ercs_total_from_all - ercs_total_from_split) > 0.01:
        warnings.append(
            f"⚠️ ERCS total in All Gifts ({ercs_total_from_all:,.2f}) does not match "
            f"sum of ERCS workbook sheets ({ercs_total_from_split:,.2f}). "
            "Check that both workbooks are from the same period."
        )

    items: list[LineItem] = []
    for label in LINE_ORDER:
        amt = round(amounts.get(label, 0.0), 2)
        if amt == 0:
            continue
        items.append(LineItem(label=label, class_num=LINE_TO_CLASS[label], amount=amt))
    for label, amt in amounts.items():
        if label not in LINE_TO_CLASS and round(amt, 2) != 0:
            warnings.append(f"⚠️ Unmapped line '{label}' with amount {amt:,.2f} — skipped.")
    return items, warnings


# ---------------------------------------------------------------------------
# IIF writer
# ---------------------------------------------------------------------------

IIF_HEADER = (
    "!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\n"
    "!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\n"
    "!ENDTRNS\t\t\t\t\t\t \t\t\t\n"
)


def _fmt_amount(x: float) -> str:
    """Match sample IIF: no fixed decimals, trailing zeros stripped."""
    s = f"{x:.3f}".rstrip("0").rstrip(".")
    return s if s else "0"


def build_iif(items: list[LineItem], posting_date: str, quarter: str, docnum: str = "In-Kind") -> str:
    out = [IIF_HEADER]
    first = True
    qtr_short = quarter.split()[0] if quarter else "Q"
    for item in items:
        class_path = CLASS_PATHS[item.class_num]
        exp_memo = f"{quarter} in-kind program exp per {qtr_short} summary - {item.label}"
        rev_memo = f"{quarter} in-kind program rev per {qtr_short} summary - {item.label}"
        tag = "TRNS" if first else "SPL"
        out.append(
            f"{tag}\t\tGENERAL JOURNAL\t{posting_date}\t{DEBIT_ACCOUNT}\t\t{class_path}\t"
            f"{_fmt_amount(item.amount)}\t{docnum if first else ''}\t{exp_memo}\n"
        )
        out.append(
            f"SPL\t\tGENERAL JOURNAL\t{posting_date}\t{CREDIT_ACCOUNT}\t\t{class_path}\t"
            f"{_fmt_amount(-item.amount)}\t\t{rev_memo}\n"
        )
        first = False
    out.append("ENDTRNS\n")
    return "".join(out)


# ---------------------------------------------------------------------------
# Excel JE writer
# ---------------------------------------------------------------------------

from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime


ACCT_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'


def build_excel_je(items: list[LineItem], posting_date: str, quarter: str) -> bytes:
    """Build an Excel General Journal matching the Cornerstones template layout."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = f"{quarter} In-Kind"

    # --- Column widths ---
    col_widths = {"A": 12, "B": 28, "C": 14, "D": 9, "E": 49, "F": 15, "G": 16}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # --- Styles ---
    bold = Font(bold=True)
    bold_center = Font(bold=True)
    center = Alignment(horizontal="center")
    medium_bottom = Border(bottom=Side(style="medium"))
    thin_bottom = Border(bottom=Side(style="thin"))

    # --- Header rows ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "Cornerstones, Inc."
    ws["A1"].font = bold
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = "General Journal"
    ws["A2"].font = bold
    ws["A2"].alignment = center

    ws["F3"] = "Gen'l Journal #"
    ws["F3"].font = bold
    ws["G3"] = "In-Kind"
    ws["G3"].border = thin_bottom

    # --- Column headers (row 5) ---
    headers = ["Posting Date", "Account Name", "Account #", "Class", "Memo", "Debit", "Credit"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = medium_bottom

    # --- Parse posting date ---
    try:
        dt = datetime.strptime(posting_date.strip(), "%m/%d/%Y")
    except ValueError:
        dt = datetime.now()

    qtr_short = quarter.split()[0] if quarter else "Q"

    # --- Data rows ---
    row = 6
    for item in items:
        exp_memo = f"{quarter} in-kind program exp per {qtr_short} summary - {item.label}"
        rev_memo = f"{quarter} in-kind program rev per {qtr_short} summary - {item.label}"

        # Debit line
        ws.cell(row=row, column=1, value=dt).number_format = "mm-dd-yy"
        ws.cell(row=row, column=2, value="In-Kind - Programs")
        ws.cell(row=row, column=3, value=9026.1)
        ws.cell(row=row, column=4, value=int(item.class_num))
        ws.cell(row=row, column=5, value=exp_memo)
        c = ws.cell(row=row, column=6, value=item.amount)
        c.number_format = ACCT_FMT
        c.font = bold
        row += 1

        # Credit line
        ws.cell(row=row, column=1, value=dt).number_format = "mm-dd-yy"
        ws.cell(row=row, column=2, value="In-Kind Goods & Prof Services")
        ws.cell(row=row, column=3, value=8111)
        ws.cell(row=row, column=4, value=int(item.class_num))
        ws.cell(row=row, column=5, value=rev_memo)
        c = ws.cell(row=row, column=7, value=item.amount)
        c.number_format = ACCT_FMT
        c.font = bold
        row += 1

    # --- Totals row ---
    last_data_row = row - 1
    total_row = row
    f_sum = ws.cell(row=total_row, column=6, value=f"=SUM(F6:F{last_data_row})")
    f_sum.number_format = ACCT_FMT
    f_sum.font = bold
    f_sum.border = thin_bottom
    g_sum = ws.cell(row=total_row, column=7, value=f"=SUM(G6:G{last_data_row})")
    g_sum.number_format = ACCT_FMT
    g_sum.font = bold
    g_sum.border = thin_bottom

    # --- Description row ---
    desc_row = total_row + 1
    ws.cell(row=desc_row, column=1, value="Description").font = bold
    ws.cell(row=desc_row, column=2, value=f"To record {quarter} gifts in kind.")

    # --- Write to bytes ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit App
# ---------------------------------------------------------------------------

def main():
    st.set_page_config(page_title="In-Kind JE → IIF Generator", page_icon="📊", layout="centered")

    st.title("📊 In-Kind JE → IIF Generator")
    st.caption("Cornerstones, Inc. — Quarterly In-Kind Gift Journal Entry")

    st.markdown("---")

    # --- File uploads ---
    st.subheader("1. Upload Source Files")

    col1, col2 = st.columns(2)
    with col1:
        all_gifts_file = st.file_uploader(
            "All Gifts workbook",
            type=["xlsx"],
            help="e.g. All Gifts Q2 FY2026 REV01.xlsx"
        )
    with col2:
        ercs_file = st.file_uploader(
            "ERCS Gifts workbook",
            type=["xlsx"],
            help="e.g. ERCS Gifts Q2 FY2026 REV01.xlsx — must have two sheets (351… and 381…)"
        )

    st.markdown("---")

    # --- Settings ---
    st.subheader("2. Journal Entry Settings")

    col_a, col_b = st.columns(2)
    with col_a:
        posting_date = st.text_input("Posting Date (MM/DD/YYYY)", value="12/31/2025")
    with col_b:
        quarter = st.text_input("Quarter Label (for memos)", value="Q2 FY26")

    st.markdown("---")

    # --- Generate ---
    st.subheader("3. Generate Files")

    if st.button("🚀 Generate IIF + Excel JE", type="primary", use_container_width=True):
        if not all_gifts_file:
            st.error("Please upload the **All Gifts** workbook.")
            return
        if not ercs_file:
            st.error("Please upload the **ERCS Gifts** workbook.")
            return
        if not posting_date.strip():
            st.error("Please enter a posting date.")
            return
        if not quarter.strip():
            st.error("Please enter a quarter label.")
            return

        with st.spinner("Reading workbooks and generating IIF..."):
            try:
                all_gifts_bytes = all_gifts_file.read()
                ercs_bytes = ercs_file.read()

                items, warnings = build_line_items(all_gifts_bytes, ercs_bytes)

                if not items:
                    st.error("No line items were generated. Check that the uploaded files are correct.")
                    return

                iif_content = build_iif(items, posting_date.strip(), quarter.strip())
                excel_bytes = build_excel_je(items, posting_date.strip(), quarter.strip())

                # Show warnings
                for w in warnings:
                    st.warning(w)

                # Success
                total = sum(i.amount for i in items)
                st.success(f"IIF generated successfully — {len(items)} program lines, total: ${total:,.2f}")

                # Summary table
                st.markdown("**Line Item Summary:**")
                table_data = []
                for item in items:
                    table_data.append({
                        "Program": item.label,
                        "Class": item.class_num,
                        "Debit": f"${item.amount:,.2f}",
                        "Credit": f"$({item.amount:,.2f})",
                    })
                table_data.append({
                    "Program": "**TOTAL**",
                    "Class": "",
                    "Debit": f"**${total:,.2f}**",
                    "Credit": f"**$({total:,.2f})**",
                })
                st.table(table_data)

                # Download buttons
                qtr_short = quarter.strip().split()[0] if quarter.strip() else "Q"

                dl_col1, dl_col2 = st.columns(2)
                with dl_col1:
                    st.download_button(
                        label="⬇️ Download IIF File",
                        data=iif_content,
                        file_name=f"In-Kind {qtr_short} IIF.iif",
                        mime="text/plain",
                        type="primary",
                        use_container_width=True,
                    )
                with dl_col2:
                    st.download_button(
                        label="⬇️ Download Excel JE",
                        data=excel_bytes,
                        file_name=f"In Kind {qtr_short} JE.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )

                # Show preview
                with st.expander("Preview IIF contents"):
                    st.code(iif_content, language=None)

            except Exception as e:
                st.error(f"Error: {e}")

    # --- Help section ---
    with st.expander("ℹ️ How it works"):
        st.markdown("""
**What this app does:**

This app reads two Excel workbooks and generates a QuickBooks IIF file
for the quarterly in-kind gift journal entry.

**Files you need to upload:**

| File | What it contains |
|------|-----------------|
| **All Gifts** workbook | Master list of all in-kind gifts with Fund Description and Gift Amount columns. Provides totals for all 8 programs. |
| **ERCS Gifts** workbook | Embry Rucker gifts split into two sheets: "351…" (Shelter) and "381…" (Homeless Households, split 50/50). |

**Why is the ERCS file needed?**

The All Gifts file shows ERCS as one lump sum, but the journal entry
needs it split into 3 lines (Shelter, HH w/Children, HH w/o Children)
with different QuickBooks classes. Only the ERCS file has that breakdown.

**Program → QuickBooks Class Mapping:**

| Program | Class Code |
|---------|-----------|
| ASAPP | 41000 |
| Community Connected Sites | 43000 |
| ESHP (HOUSE) | 51000 |
| Food Hub | 81000 |
| ERCS Shelter | 32100 |
| Hypothermia | 32200 |
| HH w/ Children | 32302 |
| HH w/o Children | 32303 |
| General CS | 21000 |
| HNRC -OAS | 42000 |
| LLC | 61000 |
        """)


if __name__ == "__main__":
    main()
